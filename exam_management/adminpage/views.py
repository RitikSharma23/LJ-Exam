from django.shortcuts import render,HttpResponse,HttpResponseRedirect,HttpResponsePermanentRedirect
from adminpage.models import *
from django.urls import reverse
from django.contrib import messages
from openpyxl import Workbook
import openpyxl
import mysql.connector
import os
import datetime
from django.http import FileResponse

institute=""
program=""

config={
'host':"localhost",
'user':"root",
'password':"",
'database':"python"
}

fields=['enrollment','sem','roll','oldenrollment','name','phone','email','gender','dob','caste','subcast','category','password','photo','institute_code','program_code','parent_contact','emergency_contact','userid','address','aadhaar','finalsem','term_end','total_credits','total_grade_points','total_backlog']

def arrange(col):
    mydb=mysql.connector.connect(**config)
    mycursor = mydb.cursor()
    sql="SELECT `COLUMN_NAME` FROM `INFORMATION_SCHEMA`.`COLUMNS` WHERE `TABLE_SCHEMA`='python' AND `TABLE_NAME`='adminpage_studentmarks';"
    mycursor.execute(sql)
    myresult = mycursor.fetchall()
    if(len(myresult)>1):
        table=list()
        for x in range(1,len(myresult)):
            table.append(myresult[x][0])

        table.append(col)
        after=int
        table.sort()
        for x in range(0,len(table)):
            if(table[x]==col):
                    after=x
        
        if(after==(len(table))-1):return("AFTER "+table[after-1])
        elif(after==0):return("AFTER enrollment")
        else:mydb.close();return("AFTER "+table[after-1])

    else: mydb.close();return('AFTER enrollment')

def clear(request):
    global institute,program
    institute=""
    program=""
    return HttpResponse("cleared")

def addcourse(request):
    global institute,program
    data=request.POST
    institute=data['institute_code']
    program=data['program_code']
    print(institute)
    return HttpResponse(data['institute_code'])

def selectcourse(request,page,data):
    global institute,program   
    db=Course.objects.values_list('institute_Code')
    if not db:
        return render(request,"addinstitute.html",{})
    else:
        if(Course.objects.values_list('institute_Code')):
            if(institute==""):
                mydata = Course.objects.values_list('institute_Code')
                details={'institute':[*set(mydata)],'pro':False,'ins':True}
                return render(request,"selectcourse.html",details)
            elif(program==""):
                mydata = Course.objects.values_list('program_code').filter(institute_Code=institute)
                details={'institut':mydata,'pro':True,'ins':False,'institute':institute}
                return render(request,"selectcourse.html",details)
            else:
                db=Course.objects.values_list('institute_Code','program_code','type','institute_Name','degree_Name','category','branch').filter(institute_Code=institute,program_code=program)
                data['course']=db[0]
                return render(request,page,data)

def adminpage(request):  
    data={}
    return selectcourse(request,"admin.html",data)

def addinstitute(request):
    data={}
    return selectcourse(request,"addinstitute.html",data)

def doaddinstitute(request):
    data=request.POST
    db=Course(institute_Code=data['institute_code'],program_code=data['program_code'],type=data['program_type'],institute_Name=data['institute_name'],degree_Name=data['degree_name'],category=data['program_category'],branch=data['branch_name'])
    x=db.save()
    return HttpResponse('{"status":"success"}')

def findinstitute(request): 
    mymembers = Course.objects.all().values()   
    data={'courselist':mymembers}
    return selectcourse(request,"findinstitute.html",data)

def editinstitute(request):
    data=request.POST
    db=Course.objects.values_list('id','institute_Code','program_code','type','institute_Name','degree_Name','category','branch').filter(id=data['uid'])
    print(db[0])
    data={
    'courselist':db[0]}
    return selectcourse(request,"editinstitute.html",data)

def deleteinstitute(request):
    data=request.POST
    member = Course.objects.get(id=data['uid'])
    member.delete()
    print(member)
    return HttpResponse('{"status":"success"}')

def doeditinstitute(request):
    data=request.POST

    v=Course.objects.filter(id=data['uid']).update(institute_Code=data['institute_code'],program_code=data['program_code'],type=data['program_type'],institute_Name=data['institute_name'],degree_Name=data['degree_name'],category=data['program_category'],branch=data['branch_name'])
    if(v==1):
        return HttpResponse('{"status":"success"}')
    else: return HttpResponse('{"status":"fail"}')


# ================  STUDENT PROFILE ==========================



def findstudent(request): 
    mymembers = StudentDetails.objects.all().filter(institute_code=institute,program_code=program).values()  
    data={'courselist':mymembers}
    return selectcourse(request,"findstudent.html",data)



def addstudent(request):
    data={}
    mydata = StudentDetails.objects.all().values()
    return selectcourse(request,"addstudent.html",data)

def doaddstudent(request):
    data=request.POST

    d=StudentDetails(enrollment=data['enrollment'],sem=data['sem'],roll=data['roll'],oldenrollment=data['oldenrollment'],name=data['name'],phone=data['phone'],email=data['email'],gender=data['gender'],dob=data['dob'],caste=data['caste'],subcast=data['subcast'],category=data['category'],password=data['password'],photo=data['photo'],institute_code=data['institute_code'],program_code=data['program_code'],parent_contact=data['parent_contact'],emergency_contact=data['emergency_contact'],userid=data['userid'],address=data['address'],aadhaar=data['aadhaar'],finalsem=data['finalsem'],term_end=data['term_end'],total_credits=data['total_credits'],total_grade_points=data['total_grade_points'],total_backlog=data['total_backlog'])
    d.save()
    c=StudentMarks(enrollment=data['enrollment'])
    c.save()

    # if(x==1):
    return HttpResponse('{"status":"success"}')
    # else: return HttpResponse('{"status":"fail"}')

def editstudent(request):
    data=request.POST
    db = StudentDetails.objects.get(enrollment=data['enrollment'])
    data={
    'studentlist':db}
    return selectcourse(request,"editstudent.html",data)

def doeditstudent(request):
    data=request.POST
    d=StudentDetails.objects.filter(enrollment=data['enrollment']).update(sem=data['sem'],roll=data['roll'],oldenrollment=data['oldenrollment'],name=data['name'],phone=data['phone'],email=data['email'],gender=data['gender'],dob=data['dob'],caste=data['caste'],subcast=data['subcast'],category=data['category'],password=data['password'],photo=data['photo'],institute_code=data['institute_code'],program_code=data['program_code'],parent_contact=data['parent_contact'],emergency_contact=data['emergency_contact'],userid=data['userid'],address=data['address'],aadhaar=data['aadhaar'],finalsem=data['finalsem'],term_end=data['term_end'],total_credits=data['total_credits'],total_grade_points=data['total_grade_points'],total_backlog=data['total_backlog'])
    return HttpResponse('{"status":"success"}')
    


#================== EXCEL =========================================


def selectexcel(request):
    data={}
    return selectcourse(request,"excelupload.html",data)

def downloadexcel(request):
    file_location = 'addstudent.xlsx'

    try:    
        with open(file_location, 'rb') as f:
           file_data = f.read() 
        response = HttpResponse(file_data, content_type='application/vnd.ms-excel')
        response['Content-Disposition'] = 'attachment; filename="addstudent.xlsx"'

    except IOError:
        response = HttpResponse('<h1>File not exist</h1>')

    return response


def generateexcel(request):

    data=request.POST
    wb = Workbook()
    dest_filename = 'addstudent.xlsx'
    ws1 = wb.active
    ws1.title = "Students Detail"
    i=1
    x=[]

    for d in data:
        x.append(data[d])

    for i in range(0,len(x)-1):
        ws1.cell(1,i+1).value=x[i]
    wb.save(filename = dest_filename) 

    wb.close()
    return HttpResponse("success")
    

def uploadexcel(request):
    global fields,institute,program
    if "GET" == request.method:
        return render(request, 'myapp/index.html', {})
    else:
        excel_file = request.FILES["excel_file"]

        wb = openpyxl.load_workbook(excel_file)

        worksheet = wb["Students Detail"]

        excel_data = list()

        for row in worksheet.iter_rows():
            row_data = list()
            for cell in row:
                    row_data.append(str(cell.value))
            if(row_data[0]!='None'):
                excel_data.append(row_data)
        

        column=list()
        query=list()
        a=dict()

        print(excel_data)
        
        for i in range(1,len(excel_data)):
            a.clear()
            for j in range(0,len(excel_data[0])):
                a[excel_data[0][j]]=excel_data[i][j]
                if(excel_data[0][j]=='phone'):a['password']=excel_data[i][j]
                if(excel_data[0][j]=='enrollment'):
                    a['photo']=excel_data[i][j]+".jpg"
                    a['roll']=excel_data[i][j][-4:]
                    c=StudentMarks(enrollment=excel_data[i][j])
                    c.save()

            a['institute_code']=institute
            a['program_code']=program
            a['total_credits']=0
            a['total_grade_points']=0
            a['total_backlog']=0
            d=StudentDetails(**a)
            d.save()
        
        detail={'excel':excel_data}
 
        return render(request,"viewexcel.html",detail)



def addsubject(request):
    data={}
    return selectcourse(request,"addsubject.html",data)

def doaddsubject(request):
    data=request.POST

    print(data)
    d=Subject(
        subjectcode=data['subjectcode'],
        sem=data['sem'],
        subjectname=data['subjectname'],
        institute_code=data['institute_code'],
        program_code=data['program_code'],
        theory=data['theory'],
        theory_marks=data['theory_marks'],
        practical=data['practical'],
        practical_marks=data['practical_marks'],
        mid=data['mid'],
        mid_marks=data['mid_marks'],
    )
    d.save()

    if(data['theory']=="True"):
        mydb=mysql.connector.connect(**config)
        mycursor = mydb.cursor()
        subt=data['sem']+"_"+data['subjectcode']+"_t"
        zt=arrange(subt)
        val="ALTER TABLE adminpage_studentmarks ADD "+ subt +" TEXT(2000) NOT NULL DEFAULT 'n' "+zt+";"
        mycursor.execute(val)
        mydb.commit()
        mydb.close()
    if(data['practical']=="True"):
        mydb=mysql.connector.connect(**config)
        mycursor = mydb.cursor()
        subp=data['sem']+"_"+data['subjectcode']+"_p"
        zp=arrange(subp)
        val="ALTER TABLE adminpage_studentmarks ADD "+ subp +" TEXT(2000) NOT NULL DEFAULT 'n' "+zp+";"
        mycursor.execute(val)
        mydb.commit()
        mydb.close()
    if(data['mid']=="True"):
        mydb=mysql.connector.connect(**config)
        mycursor = mydb.cursor()
        subm=data['sem']+"_"+data['subjectcode']+"_m"
        zm=arrange(subm)
        val="ALTER TABLE adminpage_studentmarks ADD "+ subm +" TEXT(2000) NOT NULL DEFAULT 'n' "+zm+";"
        mycursor.execute(val)
        mydb.commit()
        mydb.close()

    
    return HttpResponse('{"status":"success"}')



def viewsubject(request):
    mymembers = Subject.objects.all().filter(institute_code=institute,program_code=program).values()  
    data={'subjectlist':mymembers}
    return selectcourse(request,"findsubject.html",data)

def editsubject(request):
    data=request.POST
    mymembers = Subject.objects.all().filter(subjectcode=data['subjectcode']).values()  
    detail={'subjectlist':mymembers[0]}
    return selectcourse(request,"editsubject.html",detail)


def doeditsubject(request):
    data=request.POST

    db = Subject.objects.all().filter(subjectcode=data['subjectcode']).values() 
    db=db[0]
    print(db)
    if(db['theory']==data['theory']):print("theory")
    if(db['practical']==data['practical']):print("practical")
    if(db['mid']==data['mid']):print("mid")

    return HttpResponse('{"status":"success"}')

def upgradepage(request):
    mydb=mysql.connector.connect(**config)
    mycursor = mydb.cursor()
    mycursor.execute("SELECT DISTINCT SUBSTRING(enrollment, 1, 2) AS batch,sem as sem FROM adminpage_studentdetails WHERE institute_code='"+institute+"' and program_code='"+program+"';")

    myresult = mycursor.fetchall()
 
    print(myresult)

    data={'batchlist':myresult,
            'institute':institute,
            'program':program}
    mydb.close()
    return selectcourse(request,"upgradepage.html",data)


def upgradebatch(request):
    mydb=mysql.connector.connect(**config)
    mycursor = mydb.cursor()
    data=request.POST
    print(data)

    mycursor.execute("UPDATE adminpage_studentdetails set sem ='"+data['sem']+"' WHERE institute_code='"+data['institute_code']+"' and program_code='"+data['program_code']+"' and enrollment like '"+data['batch']+"%';")
    mydb.commit()
    mydb.close()
    x=mycursor.rowcount
    if(x>1):return HttpResponse('{"status":"success"}')
    else:return HttpResponse("failed")