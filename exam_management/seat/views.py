from django.shortcuts import render,HttpResponse
from adminpage.models import *
import mysql.connector
from openpyxl import Workbook
import openpyxl
import gc
import ast

config={
'host':"localhost",
'user':"root",
'password':"",
'database':"python"
}

# SELECT DISTINCT SUBSTRING(enrollment, 1, 2) AS ExtractString
# FROM adminpage_studentdetails;

# UPDATE adminpage_studentdetails set sem ='3' WHERE institute_code='004' and program_code='002' and enrollment like '21%';

def seathome(request):
    d=Subject.objects.all().values()
    details={'subjectlist':d}
    return render(request,"seat.html",details)

def seatoption(request):
    data=request.POST
    d=Subject.objects.filter(subjectcode=data['subjectcode']).values()
    d=d[0]

    mydb=mysql.connector.connect(**config)
    mycursor = mydb.cursor()
    mycursor.execute("SELECT DISTINCT SUBSTRING(enrollment, 1, 2) AS batch,sem as sem FROM adminpage_studentdetails WHERE institute_code='"+d['institute_code']+"' and program_code='"+d['program_code']+"';")

    myresult = mycursor.fetchall()
    mydb.close()
    details={
        'data':data,
        'subject':d,
        'batchlist':myresult,
    }
    return render(request,"seatoption.html",details)

def remedial(request):1

def adminpage(request):
    return render(request,"admin.html",{})

# =================  seat ===================

def seatsem(request):
    mydb=mysql.connector.connect(**config)
    data=request.POST
    d=Subject.objects.filter(subjectcode=data['subjectcode']).values()
    d=d[0]
  
    c=("SELECT enrollment,name FROM adminpage_studentdetails WHERE enrollment like '__"+str(d['institute_code'])+"_"+str(d['program_code'])+"%' and sem="+str(d['sem'])+";")
    mycursor = mydb.cursor()
    mycursor.execute(c)
    myresult = mycursor.fetchall()

    wb = Workbook()
    dest_filename = 'Seat.xlsx'
    ws1 = wb.active
    ws1.title = "Students Seat"
    if(len(myresult)>1):
        for i in range(0,len(myresult)):
            ws1.cell(i+1,1).value=myresult[i][0]
            ws1.cell(i+1,2).value=myresult[i][1]
            ws1.cell(i+1,3).value=(data['session']+data['year'][2:]+data['subjectcode'][1:5]+str(d['sem'])+myresult[i][0][-3:])


        wb.save(filename = dest_filename) 

        wb.close()
        mydb.close()

        
    
        return HttpResponse('{"status":"success"}')
    else:mydb.close();return HttpResponse('{"status":"nodata"}')

def downloadseat(request):
    file_location = 'Seat.xlsx'

    try:    
        with open(file_location, 'rb') as f:
           file_data = f.read() 
        response = HttpResponse(file_data, content_type='application/vnd.ms-excel')
        response['Content-Disposition'] = 'attachment; filename="sitting.xlsx"'

    except IOError:
        response = HttpResponse('<h1>File not exist</h1>')

    return response


def seatremedial(request):
    mydb=mysql.connector.connect(**config)
    data=request.POST
    d=Subject.objects.filter(subjectcode=data['subjectcode']).values()
    myresult=[]
    myresult.clear()
    d=d[0]
    b='{"Status": "F"%'

    sub=str(d['sem'])+"_"+d['subjectcode']
    # c=('SELECT  adminpage_studentmarks.enrollment,adminpage_studentdetails.name FROM adminpage_studentmarks INNER JOIN adminpage_studentdetails ON adminpage_studentdetails.enrollment=adminpage_studentmarks.enrollment WHERE '+str(sub)+' LIKE '+"'"+b+"' AND adminpage_studentmarks.enrollment LIKE '"+data['batch']+"%'" )
    # c=('''
    # SELECT adminpage_studentmarks.enrollment,adminpage_studentdetails.name,
    # adminpage_studentmarks.'''+sub+'''_t,
    # adminpage_studentmarks.'''+sub+'''_p,
    # adminpage_studentmarks.'''+sub+'''_m 
    # FROM adminpage_studentmarks INNER JOIN adminpage_studentdetails ON adminpage_studentdetails.enrollment=adminpage_studentmarks.enrollment WHERE
    # '''+sub+'''_t LIKE '{"Status": "F"%' OR 
    # '''+sub+'''_p LIKE '{"Status": "F"%' OR 
    # '''+sub+'''_m LIKE '{"Status": "F"%' AND 
    # adminpage_studentmarks.enrollment 
    # LIKE '''+"'"+data['batch']+"%'")

    c=('''
    SELECT adminpage_studentmarks.enrollment,adminpage_studentdetails.name,
    adminpage_studentmarks.'''+sub+'''_t,
    CASE WHEN adminpage_studentmarks.'''+sub+'''_t LIKE '{"Status": "F"%' THEN 'Fail'
    WHEN adminpage_studentmarks.'''+sub+'''_t = 'n' THEN 'Na' ELSE 'Pass'
    END AS Theory,
    adminpage_studentmarks.'''+sub+'''_p,
    CASE WHEN adminpage_studentmarks.'''+sub+'''_p LIKE '{"Status": "F"%' THEN 'Fail' 
    WHEN adminpage_studentmarks.'''+sub+'''_p = 'n' THEN 'Na' ELSE 'Pass'
    END AS Practical,
    adminpage_studentmarks.'''+sub+'''_m,
    CASE WHEN adminpage_studentmarks.'''+sub+'''_m LIKE '{"Status": "F"%' THEN 'Fail'
    WHEN adminpage_studentmarks.'''+sub+'''_m = 'n' THEN 'Na' ELSE 'Pass'
    END AS mid,
    adminpage_studentdetails.sem
    FROM adminpage_studentmarks INNER JOIN adminpage_studentdetails ON adminpage_studentdetails.enrollment=adminpage_studentmarks.enrollment WHERE 
    (adminpage_studentmarks.'''+sub+'''_t LIKE '{"Status": "F"%' OR
    adminpage_studentmarks.'''+sub+'''_p LIKE '{"Status": "F"%' OR
    adminpage_studentmarks.'''+sub+'''_m LIKE '{"Status": "F"%') AND
    adminpage_studentmarks.enrollment LIKE '''+"'"+data['batch']+"%'")
    mycursor = mydb.cursor()
    mycursor.execute(c)
    myresult = mycursor.fetchall()


    wb = Workbook()
    dest_filename = 'Seat.xlsx'
    ws1 = wb.active
    ws1.title = "Students Seat"
    mydb.close()

    ws1.cell(1,1).value="Enrollment"
    ws1.cell(1,2).value="Name"
    ws1.cell(1,3).value="Seat Number"
    ws1.cell(1,4).value="Theory Marks"
    ws1.cell(1,5).value="Theory Status"
    ws1.cell(1,6).value="Practical Marks"
    ws1.cell(1,7).value="Practical Status"
    ws1.cell(1,8).value="Mid Marks"
    ws1.cell(1,9).value="Mid Status"

    

    if(len(myresult)>=1):
        for i in range(0,len(myresult)):
            markst = ast.literal_eval(myresult[i][2])
            marksp = ast.literal_eval(myresult[i][4])
            marksm = ast.literal_eval(myresult[i][6])
            ws1.cell(i+2,1).value=myresult[i][0]
            ws1.cell(i+2,2).value=myresult[i][1]
            ws1.cell(i+2,3).value=(data['session']+data['year'][2:]+data['subjectcode'][1:5]+str(myresult[0][8])+str(d['sem'])+myresult[i][0][-3:])

            if str(int(data['year'])-1) in markst['year'].keys():
                ws1.cell(i+2,4).value=markst['year'][str(int(data['year'])-1)]['theory']['marks']
                ws1.cell(i+2,5).value=myresult[i][3]
            else:
                ws1.cell(i+2,4).value='Na'
                ws1.cell(i+2,5).value='Na'

            if str(int(data['year'])-1) in marksp['year'].keys():
                ws1.cell(i+2,6).value=marksp['year'][str(int(data['year'])-1)]['practical']['marks']
                ws1.cell(i+2,7).value=myresult[i][5]
            else:
                ws1.cell(i+2,6).value='Na'
                ws1.cell(i+2,7).value='Na'

            if str(int(data['year'])-1) in marksm['year'].keys():
                ws1.cell(i+2,8).value=marksm['year'][str(int(data['year'])-1)]['mid']['marks']
                ws1.cell(i+2,9).value=myresult[i][7]
            else:
                ws1.cell(i+2,8).value='Na'
                ws1.cell(i+2,9).value='Na'
            
            if str(int(data['year'])-1) not in markst['year'].keys() and str(int(data['year'])-1) not in marksp['year'].keys() and str(int(data['year'])-1) not in marksm['year'].keys():
                return HttpResponse('{"status":"nodata"}')

        wb.save(filename = dest_filename) 

        wb.close()
        return HttpResponse('{"status":"success"}')
    else:
        return HttpResponse('{"status":"nodata"}')




# SELECT adminpage_studentmarks.enrollment,adminpage_studentdetails.name,

# CASE WHEN adminpage_studentmarks.1_050020101_t LIKE '{"Status": "F"%' THEN 'Fail'
# WHEN adminpage_studentmarks.1_050020101_t = 'n' THEN 'Na' ELSE 'Pass'
# END AS Theory,

# CASE WHEN adminpage_studentmarks.1_050020101_p LIKE '{"Status": "F"%' THEN 'Fail' 
# WHEN adminpage_studentmarks.1_050020101_p = 'n' THEN 'Na' ELSE 'Pass'
# END AS Practical,

# CASE WHEN adminpage_studentmarks.1_050020101_m LIKE '{"Status": "F"%' THEN 'Fail'
# WHEN adminpage_studentmarks.1_050020101_m = 'n' THEN 'Na' ELSE 'Pass'
# END AS mid

# FROM adminpage_studentmarks INNER JOIN adminpage_studentdetails ON adminpage_studentdetails.enrollment=adminpage_studentmarks.enrollment WHERE 
# adminpage_studentmarks.enrollment LIKE '19%';